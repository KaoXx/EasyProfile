#!/usr/bin/python3

"""
Este script tiene como finalidad, la extracción de las evidencias para su posterior insercción en la plantilla de Excel, añadiendo comentarios y diversas funcinalidades
con el objetivo de ahorrar tiempo.

[!] Autor -> Jesús Andrés Altozano
[!] Fecha -> Julio 2023
[!] Version -> 0.1
[!] KPMG España


██     ██       ██████  ██    ██ ███████     ████████ ██   ██ ██ ███████     ████████  █████  ███████ ██   ██        ██  
██     ██      ██    ██ ██    ██ ██             ██    ██   ██ ██ ██             ██    ██   ██ ██      ██  ██      ██  ██ 
██     ██      ██    ██ ██    ██ █████          ██    ███████ ██ ███████        ██    ███████ ███████ █████           ██ 
██     ██      ██    ██  ██  ██  ██             ██    ██   ██ ██      ██        ██    ██   ██      ██ ██  ██      ██  ██ 
██     ███████  ██████    ████   ███████        ██    ██   ██ ██ ███████        ██    ██   ██ ███████ ██   ██        ██ 

"""
import os
import sys
from tkinter import Tk, filedialog, Label, Button, simpledialog, Toplevel,messagebox
from PIL import Image as PILImage
from PIL import ImageTk
from docx import Document
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time

excel_template_path = ""
image_filenames = []
NivelRiesgo = ""

def select_word_file():
    root = Tk()
    root.withdraw()
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Word Files", "*.docx")])
        if file_path:
            extract_images_from_word(file_path)
            add_images_to_excel(file_path)
            root.destroy()
    except ValueError:
        messagebox.showerror("Error","[!] Selecciona un fichero válido")
        


def select_excel_template():
    global excel_template_path
    root = Tk()
    root.withdraw()
    try:
        excel_template_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        root.destroy()
    except ValueError:
        messagebox.showerror("Error","[!] Selecciona un fichero válido")


def extract_images_from_word(file_path):
    try:
        global image_filenames
        doc = Document(file_path)
        images = []

        for rel in doc.part.rels.values():
            if "image" in rel.reltype:
                image_data = rel.target_part.blob
                image_filename = rel.target_ref
                images.append((image_filename, image_data))

        folder_path = os.path.splitext(file_path)[0]
        check_folder_existence(folder_path)
        os.makedirs(folder_path, exist_ok=True)

        for i, (image_filename, image_data) in enumerate(reversed(images)):
            image_path = os.path.join(folder_path, f"imagen{i+1}.png")
            with open(image_path, "wb") as f:
                f.write(image_data)
            image_filenames.append(image_filename)
    except ValueError:
        messagebox.showerror("Error","[!] Se ha producido un error al extraer las evidencias")


def add_images_to_excel(file_path):
    try:
        folder_path = os.path.splitext(file_path)[0]
        workbook = load_workbook(excel_template_path)
        sheet = workbook.active
        image_files = os.listdir(folder_path)
        y = 2

        for i, image_file in enumerate(image_files):
            image_path = os.path.join(folder_path, image_file)
            img = PILImage.open(image_path)
            img.thumbnail((600, 300))

            preview_window = Toplevel()
            preview_window.title("Previsualización de la Evidencia")

            center_window(preview_window)

            img_tk = ImageTk.PhotoImage(img)
            image_label = Label(preview_window, image=img_tk)
            image_label.pack()

            comment = add_comment(image_file)

            col_letter = 'D'

            # Insert comment to the cell
            comment_cell = sheet.cell(row=y+27, column=4)
            comment_cell.value = comment
            comment_cell.alignment = Alignment(vertical='top', wrap_text=False)

            resized_img = OpenpyxlImage(image_path)
            resized_img.width = 600
            resized_img.height = 300

            sheet.add_image(resized_img, f"{col_letter}{y+28}")
            sheet[f"{col_letter}{y+28-1}"].alignment = Alignment(vertical='top', wrap_text=False)

            y +=20
        #workbook.save(file_path)
        sheet.cell(row=10, column=4).value = generate_text()
        sheet.cell(row=26, column=4).value = generate_text3()
        sheet.cell(row=531, column=4).value = generate_text2()

        excel_file_path = os.path.join(folder_path, "Informe PwC.xlsx")
        workbook.save(excel_file_path)
        time.sleep(1)
        root.destroy()
    except ValueError:
        messagebox.showerror("Error","[!] Se ha producido un error al procesar el Excel")


def add_comment(image_file):
    try:
        comment = simpledialog.askstring("Agregar comentario", f"Ingrese un comentario para la imagen '{image_file}':")
        return comment if comment else ""
    except ValueError:
        messagebox.showerror("Error","[!] Se ha producido un error en tiempo de ejecución")


def generate_text():
    try:
        global NivelRiesgo
        text_template = 'Para la realización de la prueba del IT Rol "{Rol}", con funcionalidad de {NdR}, lo hemos probado a través del Usuario {NombreUsuario}. A continuación se muestran evidencias obtenidas en la prueba en la que se verifica los perfiles asignados al usuario dentro de la aplicación:'
        user_1 = simpledialog.askstring("Ingrese texto para 1", "Ingrese el Rol probado")
        NivelRiesgo = simpledialog.askstring("Ingrese texto para 2", "Ingrese el nivel de riesgo:")
        user_3 = simpledialog.askstring("Ingrese texto para 3", "Ingrese el nombre de usuario y id --> Nombre Apellidos (id):")

        text = text_template.format(Rol=user_1, NdR=NivelRiesgo, NombreUsuario=user_3)
        return text
    except ValueError:
        messagebox.showerror("Error","[!] Se ha producido un error en tiempo de ejecución")

def generate_text2():
    try:
        global NivelRiesgo
        text_template = 'Por lo tanto podemos llegar a la conclusión gracias a las evidencias añadidas de que el IT Rol corresponde con la funcionalidad de {NivelRiesgo}'
        text = text_template.format(NivelRiesgo=NivelRiesgo)
        return text
    except ValueError:
        messagebox.showerror("Error","[!] Se ha producido un error en tiempo de ejecución")

def generate_text3():
    try:
        global NivelRiesgo
        text_template = 'A continuación, se adjuntan las evidencias obtenidas en la prueba con el objetivo de verificar que el IT Rol tiene funcionalidades de {NivelRiesgo}, tal y como se indica:'
        text = text_template.format(NivelRiesgo=NivelRiesgo)
        return text
    except ValueError:
        messagebox.showerror("Error","[!] Se ha producido un error en tiempo de ejecución")


def center_window(window):
    # Obtener las dimensiones de la pantalla
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Calcular la posición x e y para centrar la ventana
    window_width = window.winfo_width()
    window_height = window.winfo_height()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)

    # Establecer la posición de la ventana en el centro de la pantalla
    window.geometry(f"+{x}+{y}")



def check_folder_existence(folder_path):
    if os.path.exists(folder_path):
        messagebox.showerror("Error","Ya existe una carpeta con el mismo nombre que el archivo Word seleccionado.")
        sys.exit()
    return False

def show_warning():
    toplevel = Toplevel()
    toplevel.title("Advertencia")
    label = Label(toplevel, text="Autor: Jesús Andrés Altozano \n Advertencia: Todos los derechos reservados por KPMG España. Este programa y su contenido están protegidos por las leyes de derechos de autor y propiedad intelectual.\nQueda estrictamente prohibida la reproducción, distribución o divulgación no autorizada de este programa, su código fuente y cualquier información relacionada. \nCualquier uso no autorizado está sujeto a acciones legales. Por favor, respete los derechos reservados de KPMG España.")
    label.pack()
    button = Button(toplevel, text="Continuar", command=toplevel.destroy)
    button.pack()
    center_window(toplevel)
    toplevel.wait_window()  # Esperar hasta que el usuario cierre la ventana emergente



#Flujo Inicial del programa

if __name__ == '__main__':
    try:
        root = Tk()
        root.withdraw()
        show_warning()
        root.deiconify()
        
        root.title("Easy Profile")
        center_window(root)

        label_template = Label(root, text="Seleccione una plantilla de Excel")
        label_template.pack(anchor="center", fill="x")

        button_select_template = Button(root, text="Seleccionar plantilla", command=select_excel_template)
        button_select_template.pack(anchor="center", fill="x")

        label = Label(root, text="Seleccione un archivo de Word")
        label.pack(anchor="center", fill="x")

        button_select_word = Button(root, text="Seleccionar archivo", command=select_word_file)
        button_select_word.pack(anchor="center", fill="x")
        root.mainloop()
    except ValueError:
        messagebox.showerror("Error","[!] Se ha producido un error en tiempo de ejecución")

